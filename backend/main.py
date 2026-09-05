"""Trakerz - FastAPI backend.

Fully local: SQLite file storage, no external services.
Run with: uvicorn main:app --host 127.0.0.1 --port 8000
(see ../run.sh or ../run.bat)
"""
import re
import shutil
import sqlite3
import uuid
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, List, Dict

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

import db

BASE_DIR = Path(__file__).resolve().parent.parent
FRONTEND_DIR = BASE_DIR / "frontend"
UPLOADS_DIR = BASE_DIR / "data" / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="Trakerz", version="1.0")

BILLED_STATUSES = ("invoiced", "paid")
EXPIRING_SOON_DAYS = 30
# SOW Status is a free-text, user-editable master list (Configuration > SOW
# Status), not a fixed enum - customers add their own labels (e.g. "Future
# (Upcoming)"). So "is this SOW still open" can't check for an exact
# "active" match; instead we exclude only the statuses that clearly mean the
# SOW is closed out, and treat every other status as still in play.
CLOSED_STATUSES = ("completed", "cancelled", "expired")


@app.on_event("startup")
def startup():
    db.init_db()


# ---------- Pydantic models ----------

class MilestoneIn(BaseModel):
    description: str
    amount: float = 0
    due_date: Optional[str] = None
    status: str = "pending"          # pending | invoiced | paid
    billed_date: Optional[str] = None


class SowIn(BaseModel):
    customer_id: int
    title: str
    project_title: Optional[str] = None
    project_code: Optional[str] = None
    contract_code: Optional[str] = None
    opportunity_id: Optional[str] = None
    po_number: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    total_value: float = 0
    billing_model_id: Optional[int] = None
    operating_model_id: Optional[int] = None
    status: str = "draft"            # draft | active | completed | expired | cancelled
    notes: Optional[str] = None
    doc_link: Optional[str] = None


class CustomerIn(BaseModel):
    customer_code: str
    customer_name: str
    client_partner: Optional[str] = None
    delivery_director: Optional[str] = None
    industry: Optional[str] = None
    headquarters: Optional[str] = None
    geo: Optional[str] = None


class NameIn(BaseModel):
    name: str
    details: Optional[str] = None


class ResourceIn(BaseModel):
    account_name: Optional[str] = None
    project_name: Optional[str] = None
    wbs_id: Optional[str] = None
    employee_code: Optional[str] = None
    employee_name: str
    location_id: Optional[int] = None
    employee_type_id: Optional[int] = None
    band_id: Optional[int] = None
    allocation_start_date: Optional[str] = None
    allocation_end_date: Optional[str] = None


class RevenueCellIn(BaseModel):
    sow_id: int
    fiscal_year: int
    fiscal_month: int          # 1-12, fiscal position: 1=Apr ... 9=Dec, 10=Jan, 11=Feb, 12=Mar
    projection: float = 0
    invoiced: float = 0


class RevenueSowIn(BaseModel):
    sow_id: int
    fiscal_year: int


# Revenue Management's fiscal year runs Apr-Mar. fiscal_month is a 1-12
# position within that year (not the calendar month), so this list is
# indexed the same way: FISCAL_MONTH_LABELS[0] is fiscal_month 1 (Apr).
FISCAL_MONTH_LABELS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


# ---------- helpers ----------

def _row_to_dict(row):
    return dict(row)


def _load_lookup_maps(conn):
    """Return {id: name} maps for customers, billing_models, operating_models
    so SOW rows can be annotated with human-readable names without one query
    per foreign key per row."""
    customers = {r["id"]: r["customer_name"] for r in conn.execute("SELECT id, customer_name FROM customers")}
    billing_models = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM billing_models")}
    operating_models = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM operating_models")}
    return customers, billing_models, operating_models


def _attach_names(sow: dict, customers: Dict[int, str], billing_models: Dict[int, str], operating_models: Dict[int, str]) -> dict:
    sow["customer_name"] = customers.get(sow.get("customer_id"))
    sow["billing_model_name"] = billing_models.get(sow.get("billing_model_id"))
    sow["operating_model_name"] = operating_models.get(sow.get("operating_model_id"))
    return sow


def _enrich_sow(sow: dict, milestones: List[dict]) -> dict:
    billed_total = sum(m["amount"] for m in milestones if m["status"] in BILLED_STATUSES)
    total_value = sow["total_value"] or 0
    remaining_budget = total_value - billed_total

    alerts = []
    today = date.today()
    end_date = sow.get("end_date")
    days_to_end = None
    if end_date:
        try:
            d = datetime.strptime(end_date, "%Y-%m-%d").date()
            days_to_end = (d - today).days
        except ValueError:
            days_to_end = None

    if (sow["status"] or "").strip().lower() not in CLOSED_STATUSES and days_to_end is not None:
        if days_to_end < 0:
            alerts.append("overdue")
        elif days_to_end <= EXPIRING_SOON_DAYS:
            alerts.append("expiring_soon")

    if billed_total > total_value > 0:
        alerts.append("over_budget")

    sow["billed_total"] = round(billed_total, 2)
    sow["remaining_budget"] = round(remaining_budget, 2)
    sow["days_to_end"] = days_to_end
    sow["alerts"] = alerts
    sow["milestone_count"] = len(milestones)
    return sow


def _get_sow_or_404(conn, sow_id: int) -> dict:
    row = conn.execute("SELECT * FROM sows WHERE id = ?", (sow_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="SOW not found")
    return _row_to_dict(row)


def _get_customer_or_404(conn, customer_id: int) -> dict:
    row = conn.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Customer not found")
    return _row_to_dict(row)


def _parse_iso_date(s: Optional[str]):
    """Turn a stored 'YYYY-MM-DD' string into a real date object so Excel
    treats it as a date (sortable, formattable) instead of plain text."""
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _build_workbook(sheet_title: str, headers: List[str], rows: List[list],
                     date_cols: tuple = (), currency_cols: tuple = (), widths: Optional[List[int]] = None) -> Workbook:
    """Shared .xlsx builder for the export endpoints: bold header row, plain
    data rows, optional date/currency number formatting on 1-indexed
    column numbers, and optional fixed column widths."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col in date_cols:
            cell = row[col - 1]
            if cell.value is not None:
                cell.number_format = "dd-mmm-yyyy"
        for col in currency_cols:
            row[col - 1].number_format = "#,##0.00"

    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    return wb


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _validate_sow_refs(conn, sow: SowIn):
    if not conn.execute("SELECT 1 FROM customers WHERE id = ?", (sow.customer_id,)).fetchone():
        raise HTTPException(status_code=400, detail="Selected customer does not exist")
    if sow.billing_model_id is not None and not conn.execute(
        "SELECT 1 FROM billing_models WHERE id = ?", (sow.billing_model_id,)
    ).fetchone():
        raise HTTPException(status_code=400, detail="Selected billing model does not exist")
    if sow.operating_model_id is not None and not conn.execute(
        "SELECT 1 FROM operating_models WHERE id = ?", (sow.operating_model_id,)
    ).fetchone():
        raise HTTPException(status_code=400, detail="Selected operating model does not exist")


def _get_resource_or_404(conn, resource_id: int) -> dict:
    row = conn.execute("SELECT * FROM resources WHERE id = ?", (resource_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Resource not found")
    return _row_to_dict(row)


def _load_resource_lookup_maps(conn):
    """Return {id: name} maps for the three dropdowns on the Resource
    Management form, so a list of resources can be annotated with
    human-readable names without one query per foreign key per row."""
    locations = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM locations")}
    employee_types = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM employee_types")}
    bands = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM bands")}
    return locations, employee_types, bands


def _attach_resource_names(resource: dict, locations: Dict[int, str], employee_types: Dict[int, str], bands: Dict[int, str]) -> dict:
    resource["location_name"] = locations.get(resource.get("location_id"))
    resource["employee_type_name"] = employee_types.get(resource.get("employee_type_id"))
    resource["band_name"] = bands.get(resource.get("band_id"))
    return resource


def _validate_resource_refs(conn, r: ResourceIn):
    if r.location_id is not None and not conn.execute(
        "SELECT 1 FROM locations WHERE id = ?", (r.location_id,)
    ).fetchone():
        raise HTTPException(status_code=400, detail="Selected location does not exist")
    if r.employee_type_id is not None and not conn.execute(
        "SELECT 1 FROM employee_types WHERE id = ?", (r.employee_type_id,)
    ).fetchone():
        raise HTTPException(status_code=400, detail="Selected employee type does not exist")
    if r.band_id is not None and not conn.execute(
        "SELECT 1 FROM bands WHERE id = ?", (r.band_id,)
    ).fetchone():
        raise HTTPException(status_code=400, detail="Selected band does not exist")


# ---------- SOW endpoints ----------

@app.get("/api/sows")
def list_sows(status: Optional[str] = None, customer_id: Optional[int] = None, q: Optional[str] = None):
    with db.get_db() as conn:
        rows = conn.execute("SELECT * FROM sows ORDER BY end_date IS NULL, end_date ASC").fetchall()
        sows = [_row_to_dict(r) for r in rows]
        customers, billing_models, operating_models = _load_lookup_maps(conn)
        sows = [_attach_names(s, customers, billing_models, operating_models) for s in sows]

        if status:
            sows = [s for s in sows if s["status"] == status]
        if customer_id:
            sows = [s for s in sows if s["customer_id"] == customer_id]
        if q:
            ql = q.lower()
            sows = [
                s for s in sows
                if ql in s["title"].lower() or ql in (s["customer_name"] or "").lower()
            ]

        result = []
        for s in sows:
            m_rows = conn.execute("SELECT * FROM milestones WHERE sow_id = ?", (s["id"],)).fetchall()
            milestones = [_row_to_dict(m) for m in m_rows]
            result.append(_enrich_sow(s, milestones))
        return result


@app.post("/api/sows", status_code=201)
def create_sow(sow: SowIn):
    with db.get_db() as conn:
        _validate_sow_refs(conn, sow)
        cur = conn.execute(
            """INSERT INTO sows (customer_id, title, project_title, project_code, contract_code,
               opportunity_id, po_number, start_date, end_date,
               total_value, billing_model_id, operating_model_id, status, notes, doc_link, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))""",
            (sow.customer_id, sow.title, sow.project_title, sow.project_code, sow.contract_code,
             sow.opportunity_id, sow.po_number, sow.start_date, sow.end_date,
             sow.total_value, sow.billing_model_id, sow.operating_model_id, sow.status, sow.notes, sow.doc_link),
        )
        new_id = cur.lastrowid
        row = _get_sow_or_404(conn, new_id)
        customers, billing_models, operating_models = _load_lookup_maps(conn)
        row = _attach_names(row, customers, billing_models, operating_models)
        return _enrich_sow(row, [])


@app.get("/api/sows/export")
def export_sows(status: Optional[str] = None, customer_id: Optional[int] = None, q: Optional[str] = None):
    """Export the SOW list to an .xlsx workbook. Accepts the same filters as
    GET /api/sows, so exporting from a filtered/searched view downloads
    exactly what's on screen. Registered before /api/sows/{sow_id} so the
    literal "export" path isn't swallowed by that route's int converter."""
    sows = list_sows(status=status, customer_id=customer_id, q=q)

    headers = [
        "Customer", "Title", "Project title", "Project code", "Contract code",
        "Opportunity ID", "PO#", "Start date", "End date", "TCV", "Status",
        "Billing model", "Operating model", "Document link", "Additional information",
    ]
    rows = [
        [
            s.get("customer_name") or "",
            s.get("title") or "",
            s.get("project_title") or "",
            s.get("project_code") or "",
            s.get("contract_code") or "",
            s.get("opportunity_id") or "",
            s.get("po_number") or "",
            _parse_iso_date(s.get("start_date")),
            _parse_iso_date(s.get("end_date")),
            s.get("total_value") or 0,
            s.get("status") or "",
            s.get("billing_model_name") or "",
            s.get("operating_model_name") or "",
            s.get("doc_link") or "",
            s.get("notes") or "",
        ]
        for s in sows
    ]
    date_cols = (8, 9)
    currency_cols = (10,)
    widths = [22, 28, 24, 16, 16, 16, 14, 13, 13, 14, 14, 18, 18, 30, 34]
    wb = _build_workbook("SOWs", headers, rows, date_cols=date_cols, currency_cols=currency_cols, widths=widths)
    return _xlsx_response(wb, f"trakerz_sows_{date.today().isoformat()}.xlsx")


@app.get("/api/sows/{sow_id}")
def get_sow(sow_id: int):
    with db.get_db() as conn:
        sow = _get_sow_or_404(conn, sow_id)
        customers, billing_models, operating_models = _load_lookup_maps(conn)
        sow = _attach_names(sow, customers, billing_models, operating_models)
        m_rows = conn.execute("SELECT * FROM milestones WHERE sow_id = ? ORDER BY due_date IS NULL, due_date ASC", (sow_id,)).fetchall()
        milestones = [_row_to_dict(m) for m in m_rows]
        enriched = _enrich_sow(sow, milestones)
        enriched["milestones"] = milestones
        return enriched


@app.put("/api/sows/{sow_id}")
def update_sow(sow_id: int, sow: SowIn):
    with db.get_db() as conn:
        _get_sow_or_404(conn, sow_id)
        _validate_sow_refs(conn, sow)
        conn.execute(
            """UPDATE sows SET customer_id=?, title=?, project_title=?, project_code=?, contract_code=?,
               opportunity_id=?, po_number=?, start_date=?, end_date=?,
               total_value=?, billing_model_id=?, operating_model_id=?, status=?, notes=?, doc_link=?,
               updated_at=datetime('now') WHERE id=?""",
            (sow.customer_id, sow.title, sow.project_title, sow.project_code, sow.contract_code,
             sow.opportunity_id, sow.po_number, sow.start_date, sow.end_date,
             sow.total_value, sow.billing_model_id, sow.operating_model_id, sow.status, sow.notes, sow.doc_link,
             sow_id),
        )
        row = _get_sow_or_404(conn, sow_id)
        customers, billing_models, operating_models = _load_lookup_maps(conn)
        row = _attach_names(row, customers, billing_models, operating_models)
        m_rows = conn.execute("SELECT * FROM milestones WHERE sow_id = ?", (sow_id,)).fetchall()
        return _enrich_sow(row, [_row_to_dict(m) for m in m_rows])


@app.delete("/api/sows/{sow_id}", status_code=204)
def delete_sow(sow_id: int):
    with db.get_db() as conn:
        _get_sow_or_404(conn, sow_id)
        conn.execute("DELETE FROM sows WHERE id = ?", (sow_id,))
    return None


# ---------- Milestone endpoints ----------

@app.get("/api/sows/{sow_id}/milestones")
def list_milestones(sow_id: int):
    with db.get_db() as conn:
        _get_sow_or_404(conn, sow_id)
        rows = conn.execute("SELECT * FROM milestones WHERE sow_id = ? ORDER BY due_date IS NULL, due_date ASC", (sow_id,)).fetchall()
        return [_row_to_dict(r) for r in rows]


@app.post("/api/sows/{sow_id}/milestones", status_code=201)
def create_milestone(sow_id: int, m: MilestoneIn):
    with db.get_db() as conn:
        _get_sow_or_404(conn, sow_id)
        cur = conn.execute(
            """INSERT INTO milestones (sow_id, description, amount, due_date, status, billed_date, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, datetime('now'))""",
            (sow_id, m.description, m.amount, m.due_date, m.status, m.billed_date),
        )
        row = conn.execute("SELECT * FROM milestones WHERE id = ?", (cur.lastrowid,)).fetchone()
        return _row_to_dict(row)


@app.put("/api/milestones/{milestone_id}")
def update_milestone(milestone_id: int, m: MilestoneIn):
    with db.get_db() as conn:
        existing = conn.execute("SELECT * FROM milestones WHERE id = ?", (milestone_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Milestone not found")
        conn.execute(
            """UPDATE milestones SET description=?, amount=?, due_date=?, status=?, billed_date=?,
               updated_at=datetime('now') WHERE id=?""",
            (m.description, m.amount, m.due_date, m.status, m.billed_date, milestone_id),
        )
        row = conn.execute("SELECT * FROM milestones WHERE id = ?", (milestone_id,)).fetchone()
        return _row_to_dict(row)


@app.delete("/api/milestones/{milestone_id}", status_code=204)
def delete_milestone(milestone_id: int):
    with db.get_db() as conn:
        existing = conn.execute("SELECT * FROM milestones WHERE id = ?", (milestone_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Milestone not found")
        conn.execute("DELETE FROM milestones WHERE id = ?", (milestone_id,))
    return None


# ---------- Customer endpoints (Administration > Customer Management) ----------

@app.get("/api/customers")
def list_customers(q: Optional[str] = None):
    with db.get_db() as conn:
        rows = conn.execute("SELECT * FROM customers ORDER BY customer_name COLLATE NOCASE").fetchall()
        customers = [_row_to_dict(r) for r in rows]
        if q:
            ql = q.lower()
            customers = [
                c for c in customers
                if ql in (c["customer_code"] or "").lower() or ql in (c["customer_name"] or "").lower()
            ]
        return customers


@app.post("/api/customers", status_code=201)
def create_customer(c: CustomerIn):
    with db.get_db() as conn:
        try:
            cur = conn.execute(
                """INSERT INTO customers (customer_code, customer_name, client_partner, delivery_director,
                   industry, headquarters, geo, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))""",
                (c.customer_code, c.customer_name, c.client_partner, c.delivery_director,
                 c.industry, c.headquarters, c.geo),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail=f"Customer code '{c.customer_code}' already exists")
        return _get_customer_or_404(conn, cur.lastrowid)


@app.get("/api/customers/export")
def export_customers(q: Optional[str] = None):
    """Export the customer list to an .xlsx workbook, honoring the same
    search filter as GET /api/customers. Registered before
    /api/customers/{customer_id} for the same reason as /api/sows/export."""
    customers = list_customers(q=q)

    headers = ["Customer code", "Customer name", "Client partner", "Delivery director",
               "Industry", "Headquarters", "Geo"]
    rows = [
        [
            c.get("customer_code") or "",
            c.get("customer_name") or "",
            c.get("client_partner") or "",
            c.get("delivery_director") or "",
            c.get("industry") or "",
            c.get("headquarters") or "",
            c.get("geo") or "",
        ]
        for c in customers
    ]
    widths = [18, 28, 22, 22, 20, 22, 14]
    wb = _build_workbook("Customers", headers, rows, widths=widths)
    return _xlsx_response(wb, f"trakerz_customers_{date.today().isoformat()}.xlsx")


@app.get("/api/customers/{customer_id}")
def get_customer(customer_id: int):
    with db.get_db() as conn:
        return _get_customer_or_404(conn, customer_id)


@app.put("/api/customers/{customer_id}")
def update_customer(customer_id: int, c: CustomerIn):
    with db.get_db() as conn:
        _get_customer_or_404(conn, customer_id)
        try:
            conn.execute(
                """UPDATE customers SET customer_code=?, customer_name=?, client_partner=?,
                   delivery_director=?, industry=?, headquarters=?, geo=?,
                   updated_at=datetime('now') WHERE id=?""",
                (c.customer_code, c.customer_name, c.client_partner, c.delivery_director,
                 c.industry, c.headquarters, c.geo, customer_id),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail=f"Customer code '{c.customer_code}' already exists")
        return _get_customer_or_404(conn, customer_id)


@app.delete("/api/customers/{customer_id}", status_code=204)
def delete_customer(customer_id: int):
    with db.get_db() as conn:
        _get_customer_or_404(conn, customer_id)
        conn.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
    return None


# ---------- Resource endpoints (Management > Resource Management) ----------

@app.get("/api/resources")
def list_resources(q: Optional[str] = None):
    with db.get_db() as conn:
        rows = conn.execute("SELECT * FROM resources ORDER BY employee_name COLLATE NOCASE").fetchall()
        resources = [_row_to_dict(r) for r in rows]
        locations, employee_types, bands = _load_resource_lookup_maps(conn)
        resources = [_attach_resource_names(r, locations, employee_types, bands) for r in resources]
        if q:
            ql = q.lower()
            resources = [
                r for r in resources
                if ql in (r["employee_name"] or "").lower()
                or ql in (r["employee_code"] or "").lower()
                or ql in (r["account_name"] or "").lower()
                or ql in (r["project_name"] or "").lower()
            ]
        return resources


@app.get("/api/resources/export")
def export_resources(q: Optional[str] = None):
    """Export the Resource Management list to an .xlsx workbook, honoring
    the same search filter as GET /api/resources. Registered before
    /api/resources/{resource_id} for the same route-ordering reason as the
    SOW/Customer export endpoints above."""
    resources = list_resources(q=q)

    headers = ["Account Name", "SoW Name", "WBS ID", "Employee Code", "Employee Name",
               "Location", "Employee Type", "Band", "Allocation Start Date", "Allocation End Date"]
    rows = [
        [
            r.get("account_name") or "",
            r.get("project_name") or "",
            r.get("wbs_id") or "",
            r.get("employee_code") or "",
            r.get("employee_name") or "",
            r.get("location_name") or "",
            r.get("employee_type_name") or "",
            r.get("band_name") or "",
            _parse_iso_date(r.get("allocation_start_date")),
            _parse_iso_date(r.get("allocation_end_date")),
        ]
        for r in resources
    ]
    date_cols = (9, 10)
    widths = [22, 22, 14, 16, 22, 16, 16, 12, 20, 20]
    wb = _build_workbook("Resources", headers, rows, date_cols=date_cols, widths=widths)
    return _xlsx_response(wb, f"trakerz_resources_{date.today().isoformat()}.xlsx")


@app.post("/api/resources", status_code=201)
def create_resource(r: ResourceIn):
    with db.get_db() as conn:
        _validate_resource_refs(conn, r)
        cur = conn.execute(
            """INSERT INTO resources (account_name, project_name, wbs_id, employee_code, employee_name,
               location_id, employee_type_id, band_id, allocation_start_date, allocation_end_date, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))""",
            (r.account_name, r.project_name, r.wbs_id, r.employee_code, r.employee_name,
             r.location_id, r.employee_type_id, r.band_id, r.allocation_start_date, r.allocation_end_date),
        )
        new_id = cur.lastrowid
        row = _get_resource_or_404(conn, new_id)
        locations, employee_types, bands = _load_resource_lookup_maps(conn)
        return _attach_resource_names(row, locations, employee_types, bands)


@app.get("/api/resources/{resource_id}")
def get_resource(resource_id: int):
    with db.get_db() as conn:
        row = _get_resource_or_404(conn, resource_id)
        locations, employee_types, bands = _load_resource_lookup_maps(conn)
        return _attach_resource_names(row, locations, employee_types, bands)


@app.put("/api/resources/{resource_id}")
def update_resource(resource_id: int, r: ResourceIn):
    with db.get_db() as conn:
        _get_resource_or_404(conn, resource_id)
        _validate_resource_refs(conn, r)
        conn.execute(
            """UPDATE resources SET account_name=?, project_name=?, wbs_id=?, employee_code=?, employee_name=?,
               location_id=?, employee_type_id=?, band_id=?, allocation_start_date=?, allocation_end_date=?,
               updated_at=datetime('now') WHERE id=?""",
            (r.account_name, r.project_name, r.wbs_id, r.employee_code, r.employee_name,
             r.location_id, r.employee_type_id, r.band_id, r.allocation_start_date, r.allocation_end_date,
             resource_id),
        )
        row = _get_resource_or_404(conn, resource_id)
        locations, employee_types, bands = _load_resource_lookup_maps(conn)
        return _attach_resource_names(row, locations, employee_types, bands)


@app.delete("/api/resources/{resource_id}", status_code=204)
def delete_resource(resource_id: int):
    with db.get_db() as conn:
        _get_resource_or_404(conn, resource_id)
        conn.execute("DELETE FROM resources WHERE id = ?", (resource_id,))
    return None


# ---------- Revenue Management (Management > Revenue Management) ----------
# Two views over the same underlying data. "SoW Level" is where revenue is
# actually entered - add/edit/delete a row per SOW, exactly like SOWs and
# Resources. "Account Level" is a read-only rollup of those same SOW numbers
# grouped by customer; it has no storage of its own, it's a join+sum below.

def _current_fiscal_year() -> int:
    """The fiscal year (Apr-start) that today falls in, e.g. Feb 2027 is
    still fiscal_year 2026 (the FY that started Apr 2026)."""
    today = date.today()
    return today.year if today.month >= 4 else today.year - 1


def _fiscal_months(entries: Dict[int, dict]) -> List[dict]:
    months = []
    for fm in range(1, 13):
        cell = entries.get(fm, {"projection": 0, "invoiced": 0})
        months.append({
            "fiscal_month": fm,
            "month_label": FISCAL_MONTH_LABELS[fm - 1],
            "projection": cell["projection"],
            "invoiced": cell["invoiced"],
        })
    return months


@app.get("/api/revenue/sows")
def list_revenue_sows(fiscal_year: Optional[int] = None):
    """SoW Level grid: one row per SOW explicitly added to revenue tracking
    for this fiscal year (see POST /api/revenue/sows), each with all 12
    fiscal months (Apr-Mar) - months with no entry yet default to 0 so the
    grid is ready to type into immediately after a row is added."""
    fy = fiscal_year if fiscal_year is not None else _current_fiscal_year()
    with db.get_db() as conn:
        tracked = conn.execute(
            """SELECT s.id AS sow_id, s.title AS sow_title, s.customer_id, c.customer_name,
                      bm.name AS billing_model_name
               FROM revenue_sow_accounts ra
               JOIN sows s ON s.id = ra.sow_id
               LEFT JOIN customers c ON c.id = s.customer_id
               LEFT JOIN billing_models bm ON bm.id = s.billing_model_id
               WHERE ra.fiscal_year = ?
               ORDER BY c.customer_name COLLATE NOCASE, s.title COLLATE NOCASE""",
            (fy,),
        ).fetchall()
        entries = conn.execute(
            "SELECT sow_id, fiscal_month, projection, invoiced FROM revenue_entries WHERE fiscal_year = ?",
            (fy,),
        ).fetchall()

        by_sow: Dict[int, Dict[int, dict]] = {}
        for e in entries:
            by_sow.setdefault(e["sow_id"], {})[e["fiscal_month"]] = {
                "projection": e["projection"], "invoiced": e["invoiced"],
            }

        rows = [
            {
                "sow_id": s["sow_id"],
                "sow_title": s["sow_title"],
                "customer_id": s["customer_id"],
                "customer_name": s["customer_name"] or "Unassigned",
                "billing_model_name": s["billing_model_name"],
                "months": _fiscal_months(by_sow.get(s["sow_id"], {})),
            }
            for s in tracked
        ]
        return {"fiscal_year": fy, "rows": rows}


@app.get("/api/revenue/summary")
def revenue_summary(fiscal_year: Optional[int] = None):
    """Account Level summary: read-only rollup of every tracked SOW's
    monthly numbers, grouped by customer. Nothing to add/edit/delete here -
    it's entirely derived from the SoW Level data above."""
    fy = fiscal_year if fiscal_year is not None else _current_fiscal_year()
    sow_data = list_revenue_sows(fiscal_year=fy)

    by_customer: Dict[int, dict] = {}
    order: List[int] = []
    for row in sow_data["rows"]:
        cid = row["customer_id"] or 0
        if cid not in by_customer:
            by_customer[cid] = {
                "customer_id": row["customer_id"],
                "customer_name": row["customer_name"],
                "months": [
                    {"fiscal_month": fm, "month_label": FISCAL_MONTH_LABELS[fm - 1], "projection": 0, "invoiced": 0}
                    for fm in range(1, 13)
                ],
            }
            order.append(cid)
        for i, m in enumerate(row["months"]):
            by_customer[cid]["months"][i]["projection"] += m["projection"]
            by_customer[cid]["months"][i]["invoiced"] += m["invoiced"]

    accounts = sorted((by_customer[cid] for cid in order), key=lambda a: (a["customer_name"] or "").lower())
    return {"fiscal_year": fy, "accounts": accounts}


@app.post("/api/revenue/sows")
def add_revenue_sow(payload: RevenueSowIn):
    """Add a SOW to the SoW Level Revenue Management grid for a fiscal year
    (an explicit "Add Entry" action, mirroring how SOWs/Resources are added)."""
    with db.get_db() as conn:
        sow = conn.execute(
            """SELECT s.id, s.title, s.customer_id, c.customer_name, bm.name AS billing_model_name
               FROM sows s
               LEFT JOIN customers c ON c.id = s.customer_id
               LEFT JOIN billing_models bm ON bm.id = s.billing_model_id
               WHERE s.id = ?""",
            (payload.sow_id,),
        ).fetchone()
        if not sow:
            raise HTTPException(status_code=400, detail="Selected SOW does not exist")
        conn.execute(
            "INSERT OR IGNORE INTO revenue_sow_accounts (sow_id, fiscal_year) VALUES (?, ?)",
            (payload.sow_id, payload.fiscal_year),
        )
        entries = {
            e["fiscal_month"]: e
            for e in conn.execute(
                "SELECT fiscal_month, projection, invoiced FROM revenue_entries WHERE sow_id=? AND fiscal_year=?",
                (payload.sow_id, payload.fiscal_year),
            ).fetchall()
        }
        return {
            "sow_id": sow["id"],
            "sow_title": sow["title"],
            "customer_id": sow["customer_id"],
            "customer_name": sow["customer_name"] or "Unassigned",
            "billing_model_name": sow["billing_model_name"],
            "months": _fiscal_months(entries),
        }


@app.delete("/api/revenue/sows/{sow_id}/{fiscal_year}", status_code=204)
def delete_revenue_sow(sow_id: int, fiscal_year: int):
    """Remove a SOW from the Revenue Management grid for a fiscal year,
    deleting all of its month entries for that year along with it."""
    with db.get_db() as conn:
        row = conn.execute(
            "SELECT 1 FROM revenue_sow_accounts WHERE sow_id=? AND fiscal_year=?",
            (sow_id, fiscal_year),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Revenue row not found")
        conn.execute("DELETE FROM revenue_entries WHERE sow_id=? AND fiscal_year=?", (sow_id, fiscal_year))
        conn.execute("DELETE FROM revenue_sow_accounts WHERE sow_id=? AND fiscal_year=?", (sow_id, fiscal_year))
    return None


@app.get("/api/revenue/sows/export")
def export_revenue_sows(fiscal_year: Optional[int] = None):
    fy = fiscal_year if fiscal_year is not None else _current_fiscal_year()
    data = list_revenue_sows(fiscal_year=fy)

    headers = ["Account Name", "SOW Title", "Billing Model"]
    for label in FISCAL_MONTH_LABELS:
        headers.append(f"{label} projections")
        headers.append(f"{label} invoiced")

    rows = []
    for r in data["rows"]:
        row = [r["customer_name"], r["sow_title"], r["billing_model_name"] or ""]
        for m in r["months"]:
            row.append(m["projection"])
            row.append(m["invoiced"])
        rows.append(row)

    currency_cols = tuple(range(4, len(headers) + 1))
    widths = [24, 28, 18] + [14] * (len(headers) - 3)
    wb = _build_workbook(f"Revenue SoW Level FY{fy}", headers, rows, currency_cols=currency_cols, widths=widths)
    return _xlsx_response(wb, f"trakerz_revenue_sow_level_fy{fy}_{date.today().isoformat()}.xlsx")


@app.get("/api/revenue/summary/export")
def export_revenue_summary(fiscal_year: Optional[int] = None):
    fy = fiscal_year if fiscal_year is not None else _current_fiscal_year()
    data = revenue_summary(fiscal_year=fy)

    headers = ["Account Name"]
    for label in FISCAL_MONTH_LABELS:
        headers.append(f"{label} projections")
        headers.append(f"{label} invoiced")

    rows = []
    for acc in data["accounts"]:
        row = [acc["customer_name"]]
        for m in acc["months"]:
            row.append(m["projection"])
            row.append(m["invoiced"])
        rows.append(row)

    currency_cols = tuple(range(2, len(headers) + 1))
    widths = [24] + [14] * (len(headers) - 1)
    wb = _build_workbook(f"Revenue Account Summary FY{fy}", headers, rows, currency_cols=currency_cols, widths=widths)
    return _xlsx_response(wb, f"trakerz_revenue_account_summary_fy{fy}_{date.today().isoformat()}.xlsx")


@app.put("/api/revenue/sows")
def upsert_revenue_cell(cell: RevenueCellIn):
    """Upsert one (SOW, fiscal month) cell - the inline grid calls this once
    per cell on blur rather than saving the whole grid at once."""
    if not 1 <= cell.fiscal_month <= 12:
        raise HTTPException(status_code=400, detail="fiscal_month must be between 1 and 12")
    with db.get_db() as conn:
        if not conn.execute("SELECT 1 FROM sows WHERE id = ?", (cell.sow_id,)).fetchone():
            raise HTTPException(status_code=400, detail="Selected SOW does not exist")
        # Defensive: keep the row tracked even if this cell was saved out of
        # band (e.g. a stale grid) rather than through the Add Entry flow.
        conn.execute(
            "INSERT OR IGNORE INTO revenue_sow_accounts (sow_id, fiscal_year) VALUES (?, ?)",
            (cell.sow_id, cell.fiscal_year),
        )
        conn.execute(
            """INSERT INTO revenue_entries (sow_id, fiscal_year, fiscal_month, projection, invoiced, updated_at)
               VALUES (?, ?, ?, ?, ?, datetime('now'))
               ON CONFLICT(sow_id, fiscal_year, fiscal_month)
               DO UPDATE SET projection = excluded.projection, invoiced = excluded.invoiced, updated_at = datetime('now')""",
            (cell.sow_id, cell.fiscal_year, cell.fiscal_month, cell.projection, cell.invoiced),
        )
        row = conn.execute(
            """SELECT sow_id, fiscal_year, fiscal_month, projection, invoiced FROM revenue_entries
               WHERE sow_id=? AND fiscal_year=? AND fiscal_month=?""",
            (cell.sow_id, cell.fiscal_year, cell.fiscal_month),
        ).fetchone()
        return _row_to_dict(row)


# ---------- Configuration lookups: Locations, Billing Models, Operating Models ----------
# All are simple named master lists, so they share one CRUD implementation.

def _register_lookup_crud(path: str, table: str, label: str):
    @app.get(f"/api/{path}")
    def _list(q: Optional[str] = None):
        with db.get_db() as conn:
            rows = conn.execute(f"SELECT * FROM {table} ORDER BY name COLLATE NOCASE").fetchall()
            items = [_row_to_dict(r) for r in rows]
            if q:
                ql = q.lower()
                items = [i for i in items if ql in i["name"].lower()]
            return items

    @app.post(f"/api/{path}", status_code=201)
    def _create(item: NameIn):
        with db.get_db() as conn:
            try:
                cur = conn.execute(
                    f"INSERT INTO {table} (name, details, updated_at) VALUES (?, ?, datetime('now'))",
                    (item.name, item.details),
                )
            except sqlite3.IntegrityError:
                raise HTTPException(status_code=400, detail=f"{label} '{item.name}' already exists")
            row = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (cur.lastrowid,)).fetchone()
            return _row_to_dict(row)

    @app.put(f"/api/{path}/{{item_id}}")
    def _update(item_id: int, item: NameIn):
        with db.get_db() as conn:
            existing = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (item_id,)).fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail=f"{label} not found")
            try:
                conn.execute(
                    f"UPDATE {table} SET name=?, details=?, updated_at=datetime('now') WHERE id=?",
                    (item.name, item.details, item_id),
                )
            except sqlite3.IntegrityError:
                raise HTTPException(status_code=400, detail=f"{label} '{item.name}' already exists")
            row = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (item_id,)).fetchone()
            return _row_to_dict(row)

    @app.delete(f"/api/{path}/{{item_id}}", status_code=204)
    def _delete(item_id: int):
        with db.get_db() as conn:
            existing = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (item_id,)).fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail=f"{label} not found")
            conn.execute(f"DELETE FROM {table} WHERE id = ?", (item_id,))
        return None


_register_lookup_crud("locations", "locations", "Location")
_register_lookup_crud("billing-models", "billing_models", "Billing model")
_register_lookup_crud("operating-models", "operating_models", "Operating model")
_register_lookup_crud("statuses", "statuses", "Status")
_register_lookup_crud("employee-types", "employee_types", "Employee type")
_register_lookup_crud("bands", "bands", "Band")


# ---------- File uploads (SOW documents) ----------

def _safe_filename(name: str) -> str:
    """Strip any path components and anything that isn't safe for a plain
    filename, so an uploaded file can't be used to write outside UPLOADS_DIR
    or clobber another file's extension handling."""
    name = Path(name).name
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._") or "file"
    return name


@app.post("/api/uploads", status_code=201)
async def upload_file(file: UploadFile = File(...)):
    safe_name = _safe_filename(file.filename or "file")
    stored_name = f"{uuid.uuid4().hex}_{safe_name}"
    dest = UPLOADS_DIR / stored_name
    with dest.open("wb") as out:
        shutil.copyfileobj(file.file, out)
    return {"path": f"uploads/{stored_name}", "filename": safe_name}


# ---------- Dashboard ----------

@app.get("/api/dashboard")
def dashboard():
    with db.get_db() as conn:
        sow_rows = conn.execute("SELECT * FROM sows").fetchall()
        sows = [_row_to_dict(r) for r in sow_rows]
        customers, billing_models, operating_models = _load_lookup_maps(conn)

        enriched = []
        for s in sows:
            s = _attach_names(s, customers, billing_models, operating_models)
            m_rows = conn.execute("SELECT * FROM milestones WHERE sow_id = ?", (s["id"],)).fetchall()
            enriched.append(_enrich_sow(s, [_row_to_dict(m) for m in m_rows]))

        status_counts = {}
        value_by_client = {}
        total_value = 0.0
        total_billed = 0.0
        expiring_soon = []
        overdue = []
        over_budget = []
        expiring_30 = []

        for s in enriched:
            status_counts[s["status"]] = status_counts.get(s["status"], 0) + 1
            client_label = s["customer_name"] or "Unassigned"
            value_by_client[client_label] = value_by_client.get(client_label, 0) + (s["total_value"] or 0)
            total_value += s["total_value"] or 0
            total_billed += s["billed_total"] or 0

            if "expiring_soon" in s["alerts"]:
                expiring_soon.append(s)
            if "overdue" in s["alerts"]:
                overdue.append(s)
            if "over_budget" in s["alerts"]:
                over_budget.append(s)

            # "Expiring in 30 days" card on the SOW page (open/not-yet-closed
            # SOWs only, not already past their end date).
            days_to_end = s.get("days_to_end")
            if (s["status"] or "").strip().lower() not in CLOSED_STATUSES and days_to_end is not None and 0 <= days_to_end <= 30:
                expiring_30.append({
                    "id": s["id"],
                    "title": s["title"],
                    "customer_name": s["customer_name"],
                    "end_date": s["end_date"],
                    "days_to_end": days_to_end,
                    "status": s["status"],
                    "total_value": s["total_value"],
                })

        return {
            "status_counts": status_counts,
            "value_by_client": value_by_client,
            "total_value": round(total_value, 2),
            "total_billed": round(total_billed, 2),
            "total_remaining": round(total_value - total_billed, 2),
            "sow_count": len(enriched),
            "expiring_soon": expiring_soon,
            "overdue": overdue,
            "over_budget": over_budget,
            "expiring_30": expiring_30,
        }


# ---------- Uploaded SOW documents ----------
app.mount("/uploads", StaticFiles(directory=UPLOADS_DIR), name="uploads")


# ---------- Static frontend ----------
# Explicit routes for the HTML shell itself, marked no-store so a browser
# never serves a stale cached page after the app is updated and restarted
# (the versioned ?v= query on app.js/style.css below handles the assets).
# Registered before the catch-all mount so they take precedence over it.
@app.get("/", include_in_schema=False)
def serve_index():
    return FileResponse(FRONTEND_DIR / "index.html", headers={"Cache-Control": "no-store"})


@app.get("/index.html", include_in_schema=False)
def serve_index_html():
    return FileResponse(FRONTEND_DIR / "index.html", headers={"Cache-Control": "no-store"})


# Mounted last so /api/* routes and the explicit routes above take precedence.
app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
